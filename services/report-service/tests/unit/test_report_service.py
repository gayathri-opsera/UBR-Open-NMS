"""Unit tests for report-service."""
import sys
import os

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "../../src"))

import csv
import io
from datetime import datetime, timezone

import pytest

from query_builders import (
    alarm_history_query, kpi_summary_query, inventory_query, top_alarms_query,
)
from export import (
    generate_csv, generate_xls,
    alarm_doc_to_row, kpi_doc_to_row, top_alarm_doc_to_row, inventory_doc_to_row,
    ALARM_HISTORY_HEADERS, KPI_SUMMARY_HEADERS,
)


# ── Query builder tests ───────────────────────────────────────────────────────

class TestQueryBuilders:
    def _dt(self, **kw):
        return datetime(2026, 6, 1, tzinfo=timezone.utc).replace(**kw)

    def test_alarm_history_no_scope(self):
        q = alarm_history_query({}, self._dt(), datetime.now(timezone.utc))
        assert "raisedAt" in q
        assert "$gte" in q["raisedAt"]

    def test_alarm_history_with_network_scope(self):
        q = alarm_history_query({"networkId": "net-1"}, self._dt(), datetime.now(timezone.utc))
        assert q["networkId"] == "net-1"

    def test_alarm_history_with_org_scope(self):
        q = alarm_history_query({"organizationId": "org-1"}, self._dt(), datetime.now(timezone.utc))
        assert q["organizationId"] == "org-1"

    def test_kpi_query_includes_granularity(self):
        q = kpi_summary_query({}, self._dt(), datetime.now(timezone.utc), "15MIN")
        assert q["granularity"] == "15MIN"

    def test_inventory_query_with_hierarchy_scope(self):
        q = inventory_query({"hierarchyId": "hier-1"})
        assert q["hierarchyId"] == "hier-1"

    def test_top_alarms_pipeline_structure(self):
        pipeline = top_alarms_query({}, self._dt(), datetime.now(timezone.utc))
        assert isinstance(pipeline, list)
        stages = {list(s.keys())[0] for s in pipeline}
        assert "$match" in stages
        assert "$group" in stages
        assert "$sort" in stages
        assert "$limit" in stages


# ── Export tests ──────────────────────────────────────────────────────────────

class TestExport:
    def _alarm_doc(self):
        return {
            "alarmId": "al-1", "deviceId": "dev-1",
            "severity": "MAJOR", "alarmType": "LINK_DOWN",
            "raisedAt": datetime(2026, 6, 1, 12, 0, 0, tzinfo=timezone.utc),
            "clearedAt": datetime(2026, 6, 1, 12, 5, 0, tzinfo=timezone.utc),
            "state": "CLEARED", "acknowledgedBy": "noc1",
        }

    def test_alarm_doc_to_row_includes_duration(self):
        row = alarm_doc_to_row(self._alarm_doc())
        assert row["alarmType"] == "LINK_DOWN"
        assert "300" in row["duration"]

    def test_alarm_doc_to_row_missing_clear(self):
        doc = self._alarm_doc()
        doc.pop("clearedAt")
        row = alarm_doc_to_row(doc)
        assert row["duration"] == ""

    def test_generate_csv_valid(self):
        row = alarm_doc_to_row(self._alarm_doc())
        data = generate_csv(ALARM_HISTORY_HEADERS, [row])
        reader = csv.DictReader(io.StringIO(data.decode()))
        rows = list(reader)
        assert len(rows) == 1
        assert rows[0]["alarmType"] == "LINK_DOWN"

    def test_generate_csv_empty(self):
        data = generate_csv(ALARM_HISTORY_HEADERS, [])
        lines = data.decode().strip().splitlines()
        assert len(lines) == 1
        assert "alarmId" in lines[0]

    def test_generate_xls_valid(self):
        import openpyxl
        row = alarm_doc_to_row(self._alarm_doc())
        data = generate_xls(ALARM_HISTORY_HEADERS, [row])
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws.max_row == 2  # header + 1 data row
        headers = [ws.cell(1, c).value for c in range(1, len(ALARM_HISTORY_HEADERS) + 1)]
        assert "alarmId" in headers

    def test_kpi_doc_to_row(self):
        doc = {
            "deviceId": "dev-kpi",
            "granularity": "1HOUR",
            "bucketStart": datetime(2026, 6, 1, tzinfo=timezone.utc),
            "sampleCount": 4,
            "metrics": {
                "rssi": {"avg": -65.0, "min": -70.0, "max": -60.0},
                "snr":  {"avg": 22.5, "min": 20.0, "max": 25.0},
            },
        }
        row = kpi_doc_to_row(doc)
        assert row["rssi_avg"] == "-65.0"
        assert row["snr_avg"] == "22.5"

    def test_top_alarm_doc_to_row(self):
        doc = {
            "_id": {"alarmType": "LINK_DOWN", "severity": "MAJOR"},
            "count": 42,
            "affectedDevices": ["d1", "d2", "d3"],
        }
        row = top_alarm_doc_to_row(doc)
        assert row["count"] == 42
        assert row["affectedDeviceCount"] == 3

    def test_inventory_doc_to_row(self):
        doc = {"deviceId": "d1", "deviceType": "CPE", "manufacturer": "Senao",
               "model": "ENH900EXT", "firmwareVersion": "1.2.3",
               "status": "ONLINE", "networkId": "net-1", "organizationId": "org-1"}
        row = inventory_doc_to_row(doc)
        assert row["deviceType"] == "CPE"
        assert row["firmwareVersion"] == "1.2.3"
